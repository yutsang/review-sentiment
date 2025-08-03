"""
XLSX exporter for review data
"""
import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from ..models.review import Review
from ..utils.logger import get_logger

# Import with error handling
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class XLSXExporter:
    """Excel XLSX exporter with enhanced formatting"""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.logger = get_logger()
        
        # Get output configuration
        self.output_config = config.get("output", {})
        self.output_dir = self.output_config.get("directory", "output")
        
        # Get Excel configuration
        self.excel_config = config.get("excel", {})
        self.auto_adjust = self.excel_config.get("auto_adjust_columns", True)
        self.max_width = self.excel_config.get("max_column_width", 50)
        self.sheet_name = self.excel_config.get("sheet_name", "Reviews")
        
        # Ensure output directory exists
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
    
    def export(self, reviews: List[Review], filename: str, output_dir: str = None) -> str:
        """Export reviews to XLSX format
        
        Args:
            reviews: List of Review objects to export
            filename: Base filename (without extension)
            output_dir: Output directory (optional, uses config default if not provided)
            
        Returns:
            Full path to exported file or empty string if failed
        """
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available - install with: pip install openpyxl")
            return ""
        
        if not reviews:
            self.logger.warning("No reviews to export")
            return ""
        
        # Determine output directory
        if output_dir is None:
            output_dir = self.output_dir
        
        # Ensure output directory exists
        Path(output_dir).mkdir(parents=True, exist_ok=True)
            
        # Create full file path
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        filepath = Path(output_dir) / filename
        
        try:
            # Convert reviews to DataFrame
            df = self._reviews_to_dataframe(reviews)
                
            # Export with formatting
            self._export_with_formatting(df, str(filepath))
            
            self.logger.info(f"📊 XLSX exported: {filepath} ({len(reviews)} reviews)")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Failed to export XLSX: {e}")
            return ""
    
    def _reviews_to_dataframe(self, reviews: List[Review]) -> pd.DataFrame:
        """Convert reviews to pandas DataFrame"""
        data = []
        
        for review in reviews:
            # Use the review's to_dict method if available, otherwise build manually
            if hasattr(review, 'to_dict'):
                row = review.to_dict()
            else:
                row = {
                    'platform': review.platform,
                    'review_id': review.review_id,
                    'title': review.title,
                    'content': review.content,
                    'rating': review.rating,
                    'author': review.author,
                    'date': review.date,
                    'app_version': getattr(review, 'app_version', getattr(review, 'version', '')),
                    'helpful_count': review.helpful_count,
                    'reply_count': review.reply_count,
                    'country': getattr(review, 'country', 'hk'),
                    'sort_method': getattr(review, 'sort_method', ''),
                }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _export_with_formatting(self, df: pd.DataFrame, filepath: str):
        """Export DataFrame with Excel formatting"""
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write DataFrame
            df.to_excel(writer, sheet_name=self.sheet_name, index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[self.sheet_name]
            
            # Apply formatting if openpyxl is available
            if EXCEL_AVAILABLE:
                self._apply_excel_formatting(worksheet, df)
    
    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame):
        """Apply formatting to Excel worksheet"""
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            if self.auto_adjust:
                for column in worksheet.columns:
                    length = max(len(str(cell.value or "")) for cell in column)
                    length = min(length, self.max_width)  # Cap at max_width
                    worksheet.column_dimensions[column[0].column_letter].width = length + 2
            
            # Format specific columns
            self._format_data_columns(worksheet, df)
            
        except Exception as e:
            self.logger.debug(f"Error applying Excel formatting: {e}")
    
    def _format_data_columns(self, worksheet, df: pd.DataFrame):
        """Format specific data columns"""
        try:
            # Date formatting
            if 'date' in df.columns:
                date_col = df.columns.get_loc('date') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=date_col)
                    if cell.value:
                        cell.number_format = 'YYYY-MM-DD'
            
            # Rating formatting (center align)
            if 'rating' in df.columns:
                rating_col = df.columns.get_loc('rating') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=rating_col)
                    cell.alignment = Alignment(horizontal="center")
            
            # Content formatting (wrap text)
            content_cols = ['title', 'content']
            for col_name in content_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                
        except Exception as e:
            self.logger.debug(f"Error formatting data columns: {e}") 